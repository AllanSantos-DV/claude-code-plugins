"""
rf_engine.apply — ESTÁGIO 3 (mecânico): materializa o Excel revisado.

Consome o analysis.json (já preenchido pelo agente) + o extract.json (specs das
abas) e produz o entregável:
  - copia o arquivo ORIGINAL intacto (integração invisível);
  - anexa as 13 colunas (8 validação + 5 consultivas) na(s) aba(s) de RF, com
    formatação padrão NTT, dropdowns nos enums e altura de linha dinâmica;
  - cria a aba "1. Resumo Executivo" (primeira) a partir das contagens reais;
  - cria a aba "Leyenda de Gaps" (última) a partir da lista de gaps;
  - faz scrub de termos proibidos no que escrevemos;
  - versiona a saída como _revisado_fernando_vN (nunca sobrescreve).

Uso:
  python -m rf_engine.apply _work/base/extract.json _work/base/analysis.json [-o PASTA]
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import model
from . import formatting as fmt


# enums por índice de coluna (0-based dentro das 13)
_ENUM_BY_INDEX = {
    0: model.ENUM_CRITICIDAD,
    1: model.ENUM_TIPIFICACION,
    5: model.ENUM_COMPATIBLE,
    11: model.ENUM_PRIORIDAD,
    12: model.ENUM_TIPO_MEJORA,
}
_COLORED_INDEX = {0, 11}  # criticidad e prioridad recebem cor por valor


def _spec_map(extract_data: dict) -> dict[str, dict]:
    return {s["name"]: s for s in extract_data["sheets"] if s["analyzable"]}


def _append_columns(ws, spec: dict, recs_by_row: dict[int, dict], profile) -> None:
    # Anexa DEPOIS da última coluna REALMENTE usada na aba (não só do último
    # cabeçalho): algumas abas têm conteúdo em colunas sem header (ex.: RF-EVR
    # tinha dado na col 13 sem título) — anexar pelo header sobrescreveria o
    # cliente. ws.max_column garante zero sobrescrita.
    start = max(spec["last_col"], ws.max_column or 0) + 1
    hr = spec["header_row"]

    header_font = fmt.font(bold=True, color=model.COLOR_HEADER_FG, size=9)
    header_align = fmt.align("center", "center", True)
    cell_fill = fmt.fill(model.COLOR_CELL_FILL)
    thin = fmt.border_thin()

    cols = profile.columns
    # cabeçalhos + larguras (cor do cabeçalho vem do perfil)
    for i, cdef in enumerate(cols):
        col = start + i
        c = ws.cell(row=hr, column=col, value=cdef.header)
        c.fill = fmt.fill(cdef.header_fill or model.COLOR_HEADER_BG)
        c.font = header_font
        c.alignment = header_align
        c.border = thin
        ws.column_dimensions[get_column_letter(col)].width = cdef.width

    if not recs_by_row:
        return
    data_min = hr + 1
    data_max = max(recs_by_row)

    # dropdowns nos enums definidos pelo perfil
    for i, cdef in enumerate(cols):
        if not cdef.enum:
            continue
        letter = get_column_letter(start + i)
        dv = DataValidation(type="list", formula1='"' + ",".join(cdef.enum) + '"', allow_blank=True)
        dv.error = "Valor fuera de la lista permitida"
        dv.prompt = "Seleccione un valor"
        ws.add_data_validation(dv)
        dv.add(f"{letter}{data_min}:{letter}{data_max}")

    # valores por linha
    for row_idx, rec in recs_by_row.items():
        if rec.get("na"):
            continue
        widths_for_height = []
        for i, cdef in enumerate(cols):
            col = start + i
            raw = rec.get(cdef.key, "") or ""
            val, _ = fmt.scrub_text(str(raw)) if isinstance(raw, str) else (raw, [])
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin
            cell.alignment = fmt.align("left", "top", True)
            cell.font = fmt.font(size=9)
            if cdef.colored and val in model.GRAV_FILLS:
                bg, fg = model.GRAV_FILLS[val]
                cell.fill = fmt.fill(bg)
                cell.font = fmt.font(bold=True, color=fg, size=9)
            else:
                cell.fill = cell_fill
            widths_for_height.append((val, cdef.width))
        ws.row_dimensions[row_idx].height = fmt.dynamic_row_height(widths_for_height)


def _compute_counts(records: list[dict]) -> dict:
    active = [r for r in records if not r.get("na")]
    crit = Counter((r.get("criticidad") or "").strip() for r in active if (r.get("criticidad") or "").strip())
    tip = Counter((r.get("tipificacion") or "").strip() for r in active if (r.get("tipificacion") or "").strip())
    bloqueantes = [r for r in active if (r.get("tipificacion") or "").strip().lower() == "bloqueante"]
    fuentes = []
    seen = set()
    for r in active:
        ref = (r.get("referencia") or "").strip()
        for line in ref.replace("•", "\n").split("\n"):
            line = line.strip(" -\t")
            if line and line.lower() not in seen and not line.lower().endswith(".md"):
                seen.add(line.lower())
                fuentes.append(line)
    return {"total": len(active), "crit": crit, "tip": tip,
            "bloqueantes": bloqueantes, "fuentes": fuentes[:20]}


def _build_resumo(wb, source_name: str, counts: dict) -> None:
    if "1. Resumo Executivo" in wb.sheetnames:
        wb.remove(wb["1. Resumo Executivo"])
    ws = wb.create_sheet("1. Resumo Executivo", 0)
    for col, w in zip("ABCDEFGH", [28, 14, 18, 55, 4, 4, 4, 4]):
        ws.column_dimensions[col].width = w

    def band(row, text, bg, size=11, height=None):
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.fill = fmt.fill(bg)
        c.font = fmt.font(bold=True, color="FFFFFF", size=size)
        c.alignment = fmt.align("left", "center", True)
        if height:
            ws.row_dimensions[row].height = height

    band(1, "PROYECTO ONE - LA POSITIVA SEGUROS", "0C447C", 12, 21.75)
    band(2, f"Resumen Ejecutivo\n{source_name}", "185FA5", 11, 31.5)

    meta = [("Cliente:", "La Positiva Seguros"), ("Proyecto:", "Proyecto ONE - InsureMO"),
            ("Version:", "1.0"), ("Preparado por:", "Fernando Soares - InsureMO")]
    r = 4
    for label, value in meta:
        cl = ws.cell(row=r, column=1, value=label)
        cl.fill = fmt.fill("D6E4F0")
        cl.font = fmt.font(bold=True, size=9)
        ws.merge_cells(f"B{r}:H{r}")
        ws.cell(row=r, column=2, value=value).font = fmt.font(size=9)
        r += 1

    r += 1
    band(r, "Distribución por Gravedad / Atención", "185FA5", 10, 18)
    r += 1
    hdr = ["Gravedad", "Color", "Cantidad", "Descripción"]
    for j, h in enumerate(hdr):
        c = ws.cell(row=r, column=1 + j, value=h)
        c.fill = fmt.fill("0C447C")
        c.font = fmt.font(bold=True, color="FFFFFF", size=9)
        c.alignment = fmt.align("center", "center", True)
    ws.merge_cells(f"D{r}:H{r}")
    r += 1
    descr = {"Alta": "Bloquea entrega / riesgo alto", "Media": "Requiere ajuste o validación",
             "Baja": "Confirmado / menor", "Fuera de alcance": "Fuera de Fase I"}
    for grav in ["Alta", "Media", "Baja", "Fuera de alcance"]:
        bg, fg = model.GRAV_FILLS.get(grav, ("FFFFFF", "000000"))
        ws.cell(row=r, column=1, value=grav).font = fmt.font(size=9)
        sw = ws.cell(row=r, column=2, value="")
        sw.fill = fmt.fill(model.SWATCHES.get(grav, "FFFFFF"))
        ws.cell(row=r, column=3, value=counts["crit"].get(grav, 0)).font = fmt.font(bold=True, size=9)
        ws.merge_cells(f"D{r}:H{r}")
        dc = ws.cell(row=r, column=4, value=descr[grav])
        dc.fill = fmt.fill(bg)
        dc.font = fmt.font(color=fg, size=9)
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = fmt.font(bold=True, color="FFFFFF", size=9)
    ws.cell(row=r, column=1).fill = fmt.fill("0C447C")
    ws.cell(row=r, column=3, value=counts["total"]).font = fmt.font(bold=True, size=9)
    r += 2

    band(r, "Puntos para revisar antes del envío", "185FA5", 10, 18)
    r += 1
    if counts["bloqueantes"]:
        for b in counts["bloqueantes"][:15]:
            txt = f"• [{b.get('rf_id','')}] {(b.get('comentario') or b.get('resumen_funcional') or '').strip()}"
            ws.merge_cells(f"A{r}:H{r}")
            c = ws.cell(row=r, column=1, value=fmt.scrub_text(txt)[0])
            c.alignment = fmt.align("left", "top", True)
            c.font = fmt.font(size=9)
            r += 1
    else:
        ws.merge_cells(f"A{r}:H{r}")
        ws.cell(row=r, column=1, value="• Sin bloqueantes registrados.").font = fmt.font(size=9)
        r += 1

    r += 1
    band(r, "Fuentes usadas en el análisis", "185FA5", 10, 18)
    r += 1
    for fsrc in (counts["fuentes"] or ["• (referencias se completan al llenar el análisis)"]):
        ws.merge_cells(f"A{r}:H{r}")
        c = ws.cell(row=r, column=1, value=("• " + fsrc) if not fsrc.startswith("•") else fsrc)
        c.alignment = fmt.align("left", "top", True)
        c.font = fmt.font(size=9)
        r += 1


def _build_leyenda(wb, gaps: list[dict]) -> None:
    if "Leyenda de Gaps" in wb.sheetnames:
        wb.remove(wb["Leyenda de Gaps"])
    ws = wb.create_sheet("Leyenda de Gaps")
    cols = ["Código", "Descripción corta", "Acción esperada", "Responsable sugerido", "RF relacionado"]
    for col, w in zip("ABCDE", [16, 45, 45, 22, 24]):
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value="Leyenda de Gaps")
    t.fill = fmt.fill("0C447C")
    t.font = fmt.font(bold=True, color="FFFFFF", size=12)
    for j, h in enumerate(cols):
        c = ws.cell(row=2, column=1 + j, value=h)
        c.fill = fmt.fill("0C447C")
        c.font = fmt.font(bold=True, color="FFFFFF", size=9)
        c.alignment = fmt.align("center", "center", True)
    r = 3
    for g in gaps:
        vals = [g.get("codigo", ""), g.get("descripcion", ""), g.get("accion_esperada", ""),
                g.get("responsable", ""), g.get("rf_relacionado", "")]
        crit = (g.get("criticidad") or "Media").strip()
        bg, fg = model.GRAV_FILLS.get(crit, ("FFFFFF", "000000"))
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + j, value=fmt.scrub_text(str(v))[0])
            c.alignment = fmt.align("left", "top", True)
            c.font = fmt.font(size=9)
            if j == 0:
                c.fill = fmt.fill(bg)
                c.font = fmt.font(bold=True, color=fg, size=9)
        r += 1
    if not gaps:
        ws.merge_cells(f"A{r}:E{r}")
        ws.cell(row=r, column=1, value="(sin gaps registrados)").font = fmt.font(size=9)


def _version_out(source_file: str, out_dir: str | None) -> Path:
    src = Path(source_file)
    base = src.stem
    # remove sufixo _CT/_V\d/ fernando etc. mantendo o núcleo do nome do cliente
    outp = Path(out_dir) if out_dir else src.parent
    outp.mkdir(parents=True, exist_ok=True)
    n = 1
    while True:
        cand = outp / f"{base}_revisado_fernando_v{n}.xlsx"
        if not cand.exists():
            return cand
        n += 1


def run(extract_json: str, analysis_json: str, out_dir: str | None = None,
        profile_id: str | None = None) -> dict:
    from . import profiles
    ex = json.loads(Path(extract_json).read_text(encoding="utf-8"))
    an = json.loads(Path(analysis_json).read_text(encoding="utf-8"))
    source = ex["source_file"]
    if not Path(source).exists():
        sys.exit(f"ERRO: arquivo original não encontrado: {source}")

    # perfil: argumento > gravado no analysis.json > rf-end
    profile = profiles.get_profile(profile_id or an.get("profile") or "rf-end")

    specs = _spec_map(ex)
    recs_by_sheet: dict[str, dict[int, dict]] = defaultdict(dict)
    for rec in an["records"]:
        recs_by_sheet[rec["sheet"]][rec["row_idx"]] = rec

    wb = openpyxl.load_workbook(source)  # data_only=False -> preserva fórmulas/estilos
    touched = []
    for name, spec in specs.items():
        if name in wb.sheetnames:
            _append_columns(wb[name], spec, recs_by_sheet.get(name, {}), profile)
            touched.append(name)

    counts = _compute_counts(an["records"])
    # Resumo/Leyenda só quando o perfil pede (rf-end tem criticidad; perfis autorais
    # do cliente podem não ter — nesse caso não injetamos abas nossas)
    if profile.build_resumo:
        _build_resumo(wb, ex["source_name"], counts)
    if profile.build_leyenda:
        _build_leyenda(wb, an.get("gaps", []))

    out_path = _version_out(source, out_dir)
    if str(out_path.resolve()).lower() == str(Path(source).resolve()).lower():
        sys.exit("ERRO: saída não pode ser o arquivo original.")
    wb.save(out_path)

    return {"output": str(out_path), "sheets_touched": touched, "profile": profile.id, "counts": {
        "total": counts["total"], "bloqueantes": len(counts["bloqueantes"]),
        "gaps": len(an.get("gaps", [])), "crit": dict(counts["crit"])}}


def main() -> None:
    p = argparse.ArgumentParser(description="Materializa o Excel revisado (estágio 3 do motor).")
    p.add_argument("extract", help="extract.json (estágio 1)")
    p.add_argument("analysis", help="analysis.json preenchido (agente)")
    p.add_argument("-o", "--out", default=None, help="pasta de saída (default: junto do original)")
    p.add_argument("--perfil", default=None, help="id do perfil de colunas (default: rf-end)")
    args = p.parse_args()
    res = run(args.extract, args.analysis, args.out, args.perfil)
    print("OK apply:")
    print(f"  saída: {res['output']}")
    print(f"  perfil: {res['profile']} | abas anotadas: {res['sheets_touched']}")
    print(f"  totais: {res['counts']}")


if __name__ == "__main__":
    main()
