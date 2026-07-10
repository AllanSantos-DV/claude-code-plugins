"""
rf_engine.extract — ESTÁGIO 1 (mecânico): lê o Excel e emite estrutura limpa.

O agente nunca processa o .xlsx cru. Esta tool:
  1. autodetecta as abas analisáveis pela ESTRUTURA (não pelo nome);
  2. detecta a linha de cabeçalho por aba (Listado=4, sub-fluxos=1);
  3. resolve a sentinela do requisito (col A com fórmula -> usa col B);
  4. extrai cada linha (rf_id + campos) e os hyperlinks presentes;
  5. grava extract.json (consumido pelo scaffold) + um .md limpo por aba.

Uso:
  python -m rf_engine.extract ARQUIVO.xlsx [-o PASTA_SAIDA]
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl

from . import model


def _norm(v) -> str:
    return "" if v is None else str(v).strip()


def detect_header_row(ws) -> tuple[int, int]:
    """Retorna (linha_cabecalho, score) pela presença das assinaturas de RF."""
    best = (1, -1)
    scan = min(model.HEADER_SCAN_ROWS, ws.max_row or 1)
    for r in range(1, scan + 1):
        vals = [_norm(ws.cell(r, c).value).lower() for c in range(1, min(ws.max_column or 1, 45) + 1)]
        score = sum(any(sig in v for v in vals) for sig in model.RF_HEADER_SIGNATURES)
        if score > best[1]:
            best = (r, score)
    return best


def _last_nonempty_col(ws, header_row: int) -> int:
    last = 0
    for c in range(1, (ws.max_column or 1) + 1):
        if _norm(ws.cell(header_row, c).value):
            last = c
    return last


def _find_col(headers: dict[int, str], *needles: str) -> int | None:
    for c, h in headers.items():
        low = h.lower()
        if any(n in low for n in needles):
            return c
    return None


def _resolve_id_col(ws, header_row: int, columns: list[str]) -> int:
    """Resolve a coluna do ID do requisito.

    Prioridade:
      1. cabeçalho literal "Id" (ou "id rf"/"rf id"/"id requerimiento");
      2. sentinela de fórmula: 1ª coluna de dados vazia mas a 2ª preenchida -> 2;
      3. fallback: coluna 1.
    Evita casar com "Id proceso" / "Id Miro" (que não são o ID do requisito).
    """
    exact = {"id", "id rf", "rf id", "id requerimiento", "id req", "id_rf"}
    for c, h in enumerate(columns, start=1):
        low = h.strip().lower()
        if low in exact:
            return c
    # heurística de fórmula (col A com CONCAT retorna None em data_only)
    col_a_has = col_b_has = 0
    for r in range(header_row + 1, min(header_row + 4, (ws.max_row or header_row) + 1)):
        if _norm(ws.cell(r, 1).value):
            col_a_has += 1
        if ws.max_column and ws.max_column >= 2 and _norm(ws.cell(r, 2).value):
            col_b_has += 1
    if col_a_has == 0 and col_b_has > 0:
        return 2
    return 1


def build_sheet_spec(ws) -> model.SheetSpec:
    header_row, score = detect_header_row(ws)
    analyzable = score >= model.ANALYZABLE_MIN_SIGNATURE
    last_col = _last_nonempty_col(ws, header_row) if analyzable else (ws.max_column or 0)
    columns = [_norm(ws.cell(header_row, c).value) for c in range(1, last_col + 1)] if analyzable else []
    id_col = _resolve_id_col(ws, header_row, columns) if analyzable else 1
    n_rows = 0
    if analyzable:
        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            if any(_norm(ws.cell(r, c).value) for c in range(1, last_col + 1)):
                n_rows += 1
    return model.SheetSpec(
        name=ws.title, analyzable=analyzable, header_row=header_row,
        id_col=id_col, last_col=last_col, columns=columns, n_data_rows=n_rows,
    )


def extract_rows(ws, spec: model.SheetSpec) -> list[model.RowRef]:
    headers = {c: spec.columns[c - 1] for c in range(1, spec.last_col + 1)}
    col_tipo = _find_col(headers, "tipo")
    col_imo = _find_col(headers, "insuremo", "insuremo")
    col_compl = _find_col(headers, "complej", "complex")
    col_alc = _find_col(headers, "alcance")
    col_desc = _find_col(headers, "descrip", "requerimiento")

    rows: list[model.RowRef] = []
    for r in range(spec.header_row + 1, (ws.max_row or spec.header_row) + 1):
        cells = {headers[c]: _norm(ws.cell(r, c).value) for c in range(1, spec.last_col + 1)}
        if not any(cells.values()):
            continue
        rf_id = _norm(ws.cell(r, spec.id_col).value) or _norm(ws.cell(r, 1).value) or _norm(ws.cell(r, 2).value)

        tipo = cells.get(headers.get(col_tipo, ""), "") if col_tipo else ""
        imo = cells.get(headers.get(col_imo, ""), "") if col_imo else ""
        compl = cells.get(headers.get(col_compl, ""), "") if col_compl else ""
        alc = cells.get(headers.get(col_alc, ""), "") if col_alc else ""
        desc = cells.get(headers.get(col_desc, ""), "") if col_desc else ""

        # heurística N/A: sem descrição real -> linha não anotável
        is_na = (not desc) and (not rf_id)
        # hint determinístico de bloqueante: escopo aberto + alta complexidade
        forced = (
            ("por confirmar" in (alc + imo).lower() or "por definir" in (alc + imo).lower())
            and ("alta" in compl.lower() or "por definir" in compl.lower())
        )

        # hyperlinks presentes na linha
        links = []
        for c in range(1, spec.last_col + 1):
            hl = ws.cell(r, c).hyperlink
            if hl and getattr(hl, "target", None):
                links.append(hl.target)

        fields = dict(cells)
        if links:
            fields["_hyperlinks"] = links
        rows.append(model.RowRef(
            sheet=spec.name, row_idx=r, rf_id=rf_id, fields=fields,
            is_na=is_na, forced_bloqueante=forced,
        ))
    return rows


def _md_table(headers: list[str], rows: list[list[str]], cap: int = 60) -> str:
    def clip(x: str) -> str:
        x = (x or "").replace("\n", " ").replace("|", "/")
        return x if len(x) <= cap else x[: cap - 1] + "…"
    head = "| " + " | ".join(clip(h) for h in headers) + " |"
    sep = "| " + " | ".join("---" for _ in headers) + " |"
    body = "\n".join("| " + " | ".join(clip(c) for c in row) + " |" for row in rows)
    return f"{head}\n{sep}\n{body}\n"


def run(xlsx_path: str, out_dir: str) -> dict:
    src = Path(xlsx_path)
    if not src.exists():
        sys.exit(f"ERRO: arquivo não encontrado: {src}")
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    (out / "md").mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(src, data_only=True)
    sheets_out = []
    md_index = []
    for name in wb.sheetnames:
        ws = wb[name]
        spec = build_sheet_spec(ws)
        entry = {
            "name": spec.name, "analyzable": spec.analyzable, "header_row": spec.header_row,
            "id_col": spec.id_col, "last_col": spec.last_col, "columns": spec.columns,
            "n_data_rows": spec.n_data_rows, "rows": [],
        }
        if spec.analyzable:
            rows = extract_rows(ws, spec)
            entry["rows"] = [
                {"row_idx": rr.row_idx, "rf_id": rr.rf_id, "is_na": rr.is_na,
                 "forced_bloqueante": rr.forced_bloqueante, "fields": rr.fields}
                for rr in rows
            ]
            # md limpo por aba
            md_headers = spec.columns
            md_rows = [[rr.fields.get(h, "") for h in md_headers] for rr in rows]
            (out / "md" / f"{spec.name}.md").write_text(
                f"## {spec.name}\n_{len(rows)} linhas × {len(md_headers)} colunas (cabeçalho linha {spec.header_row})_\n\n"
                + _md_table(md_headers, md_rows),
                encoding="utf-8",
            )
            md_index.append(f"- **{spec.name}**: {len(rows)} linhas (analisável)")
        else:
            md_index.append(f"- {spec.name}: não analisável (intocável)")
        sheets_out.append(entry)
    wb.close()

    result = {
        "source_file": str(src),
        "source_name": src.name,
        "n_sheets": len(sheets_out),
        "analyzable_sheets": [s["name"] for s in sheets_out if s["analyzable"]],
        "sheets": sheets_out,
    }
    (out / "extract.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    (out / "md" / "_index.md").write_text(
        f"# Extração — {src.name}\n\nAbas ({len(sheets_out)}):\n" + "\n".join(md_index) + "\n",
        encoding="utf-8",
    )
    return result


def main() -> None:
    p = argparse.ArgumentParser(description="Extrai estrutura + linhas de um Excel de RF (estágio 1 do motor).")
    p.add_argument("xlsx", help="Arquivo .xlsx de entrada")
    p.add_argument("-o", "--out", default="_work", help="Pasta de saída (default: _work)")
    args = p.parse_args()
    res = run(args.xlsx, args.out)
    print(f"OK extract: {res['source_name']}")
    print(f"  abas: {res['n_sheets']} | analisáveis: {len(res['analyzable_sheets'])} -> {res['analyzable_sheets']}")
    tot = sum(len(s['rows']) for s in res['sheets'])
    print(f"  linhas extraídas: {tot}")
    print(f"  saída: {Path(args.out).resolve()}\\extract.json (+ md/)")


if __name__ == "__main__":
    main()
