"""
rf_engine.validate — ESTÁGIO 4 (gate): valida a análise ANTES/DEPOIS de materializar.

Espelha o gate real do Fernando (v9-quality-gate.md + validate_review_workbook.py +
anti-padrões da skill analisar-cruzado). Opera sobre o analysis.json (conteúdo) e,
opcionalmente, sobre o .xlsx de saída (erros de fórmula, termos proibidos, estrutura).

ERROS bloqueiam a entrega; WARNINGS são avisos. Exit code 1 se houver ERROS.

Uso:
  python -m rf_engine.validate _work/base/analysis.filled.json [--xlsx saida.xlsx] [--extract extract.json]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

from . import model

FORMULA_ERRORS = ["#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A"]
FORBIDDEN_REF = [
    ".md", "anotado", "claude", "codex", "chatgpt", "openai", "analysis_",
    "memoria:", "cache\\", "_revisado_fernando_v", "workbook",
    "documentacion funcional y tecnica", "documentación funcional y técnica",
]
FORBIDDEN_ANY = [
    "codex", "chatgpt", "openai", "inteligencia artificial", "inteligência artificial",
    "modelo de ia", "generado por ia", "claude", " gpt",
]


def _active(records: list[dict]) -> list[dict]:
    return [r for r in records if not r.get("na")]


def _validate_generic(data: dict, recs: list[dict]) -> tuple[list[str], list[str]]:
    """Checagens agnósticas de perfil: campos obrigatórios do perfil não-vazios,
    enums do perfil respeitados, termos proibidos, template rotation na maior coluna
    de texto. Usado por perfis que não são o rf-end."""
    from . import profiles
    errors: list[str] = []
    warns: list[str] = []
    prof = profiles.get_profile(data.get("profile") or "rf-end")

    def loc(r):
        return f"{r.get('sheet','')}!{r.get('rf_id','') or r.get('row_idx','')}"

    text_key = None
    text_w = -1
    for cdef in prof.columns:
        if not cdef.enum and cdef.width > text_w:
            text_w, text_key = cdef.width, cdef.key

    for r in recs:
        for cdef in prof.columns:
            v = (r.get(cdef.key) or "").strip()
            if not v:
                errors.append(f"{loc(r)}: coluna '{cdef.header}' vazia.")
                break
            if cdef.enum and v not in cdef.enum:
                errors.append(f"{loc(r)}: '{cdef.header}'='{v}' fora do enum {cdef.enum}.")
        blob = " ".join(str(r.get(c.key, "")) for c in prof.columns).lower()
        for bad in FORBIDDEN_ANY:
            if bad in blob:
                errors.append(f"{loc(r)}: término prohibido '{bad.strip()}'.")
                break

    if text_key:
        texts = [(r.get(text_key) or "").strip() for r in recs if (r.get(text_key) or "").strip()]
        if texts:
            common, n = Counter(texts).most_common(1)[0]
            if n > max(3, 0.3 * len(texts)):
                errors.append(f"Template rotation em '{text_key}': {n}/{len(texts)} idénticos.")
    return errors, warns


def validate_analysis(data: dict) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warns: list[str] = []
    recs = _active(data.get("records", []))
    if not recs:
        warns.append("Nenhum registro ativo (todos N/A?).")
        return errors, warns

    # perfis autorais (não rf-end) usam o gate genérico
    if (data.get("profile") or "rf-end") != "rf-end":
        return _validate_generic(data, recs)

    def loc(r):
        return f"{r.get('sheet','')}!{r.get('rf_id','') or r.get('row_idx','')}"

    # E1 — zero bloqueante no workbook inteiro
    has_bloq = any(
        (r.get("tipificacion", "").strip().lower() == "bloqueante")
        or ("bloqueante" in (r.get("comentario", "") or "").lower())
        for r in recs
    )
    if not has_bloq:
        errors.append("Workbook sem NENHUM bloqueante — revisar critério de severidade (anti-padrão #3).")

    # enums + campos obrigatórios + referência + ação + resumen==requisito
    for r in recs:
        crit = (r.get("criticidad") or "").strip()
        tip = (r.get("tipificacion") or "").strip()
        comp = (r.get("compatible") or "").strip()
        pri = (r.get("prioridad") or "").strip()
        tme = (r.get("tipo_mejora") or "").strip()
        comentario = (r.get("comentario") or "").strip()
        referencia = (r.get("referencia") or "").strip()
        accion = (r.get("accion") or "").strip()
        resumen = (r.get("resumen_funcional") or "").strip()

        if crit and crit not in model.ENUM_CRITICIDAD:
            errors.append(f"{loc(r)}: criticidad inválida '{crit}'.")
        if tip and tip not in model.ENUM_TIPIFICACION:
            errors.append(f"{loc(r)}: tipificacion inválida '{tip}'.")
        if comp and comp not in model.ENUM_COMPATIBLE:
            errors.append(f"{loc(r)}: compatible inválido '{comp}'.")
        if pri and pri not in model.ENUM_PRIORIDAD:
            errors.append(f"{loc(r)}: prioridad inválida '{pri}'.")
        if tme and tme not in model.ENUM_TIPO_MEJORA:
            errors.append(f"{loc(r)}: tipo_mejora inválido '{tme}'.")

        # obrigatórios em linha de requisito
        if not comentario:
            errors.append(f"{loc(r)}: comentario vazio.")
        if not referencia:
            errors.append(f"{loc(r)}: referencia vazia.")
        if not accion:
            errors.append(f"{loc(r)}: acción vacía.")
        elif "responsable sugerido" not in accion.lower():
            errors.append(f"{loc(r)}: 'Acción a tomar' sin 'Responsable sugerido'.")

        # E3 — referência autorreferente / proibida
        low_ref = referencia.lower()
        for bad in FORBIDDEN_REF:
            if bad in low_ref:
                errors.append(f"{loc(r)}: referencia proibida contém '{bad}' -> {referencia[:60]!r}")
                break

        # E5 — resumen == requisito (cópia)
        req = (r.get("hint", {}).get("descripcion") or r.get("hint", {}).get("requerimiento") or "").strip()
        if req and resumen and resumen == req:
            errors.append(f"{loc(r)}: resumen_funcional é cópia do requisito (anti-padrão #4).")

        # forced bloqueante deve virar bloqueante
        if r.get("hint", {}).get("forced_bloqueante") and tip.lower() != "bloqueante":
            errors.append(f"{loc(r)}: hint forced_bloqueante=true mas tipificacion='{tip}' (deveria ser Bloqueante).")

        # 5 consultivas obrigatórias
        for key in model.CONSULT_KEYS:
            if not (r.get(key) or "").strip():
                errors.append(f"{loc(r)}: columna consultiva '{key}' vacía.")
                break

        # termos proibidos em qualquer campo
        blob = " ".join(str(r.get(k, "")) for k in model.ALL_ANALYSIS_KEYS).lower()
        for bad in FORBIDDEN_ANY:
            if bad in blob:
                errors.append(f"{loc(r)}: término prohibido '{bad.strip()}' en columnas de análisis.")
                break

    # E6 — template rotation no comentario (>30% idênticos)
    comentarios = [(r.get("comentario") or "").strip() for r in recs if (r.get("comentario") or "").strip()]
    if comentarios:
        common, n = Counter(comentarios).most_common(1)[0]
        if n > max(3, 0.3 * len(comentarios)):
            errors.append(f"Template rotation: {n}/{len(comentarios)} comentarios idénticos.")

    # W — diversidade de referencia / obs_tecnica
    refs = [(r.get("referencia") or "").strip() for r in recs if (r.get("referencia") or "").strip()]
    if refs:
        common, n = Counter(refs).most_common(1)[0]
        if n > max(3, 0.5 * len(refs)):
            warns.append(f"Referencia poco diversa: {n}/{len(refs)} filas comparten la misma referencia.")
    obs = [(r.get("obs_tecnica") or "").strip() for r in recs if (r.get("obs_tecnica") or "").strip()]
    if obs:
        common, n = Counter(obs).most_common(1)[0]
        if n > max(3, 0.3 * len(obs)):
            warns.append(f"Observación técnica poco diversa: {n}/{len(obs)} filas iguales.")

    # W — distribuição de responsável 100% um só
    owners = []
    for r in recs:
        m = re.search(r"responsable sugerido:\s*([^\n]+)", (r.get("accion") or ""), re.IGNORECASE)
        if m:
            owners.append(m.group(1).strip())
    if owners:
        common, n = Counter(owners).most_common(1)[0]
        if n == len(owners) and len(owners) > 5:
            warns.append(f"Distribución de responsable 100% '{common}' — revisar.")

    # gaps referenciados existem?
    gap_codes = {g.get("codigo", "").strip() for g in data.get("gaps", [])}
    used = set()
    for r in recs:
        for m in re.findall(r"G-[A-Z]{1,6}-\d{1,3}", (r.get("comentario", "") + " " + r.get("referencia", ""))):
            used.add(m)
    missing = used - gap_codes
    if missing:
        errors.append(f"Códigos de gap usados sem entrada na Leyenda: {sorted(missing)}")

    return errors, warns


def validate_xlsx(xlsx_path: str, extract_json: str | None, profile_id: str | None = None) -> tuple[list[str], list[str]]:
    import openpyxl
    import warnings as _w
    from . import profiles
    _w.filterwarnings("ignore")
    errors: list[str] = []
    warns: list[str] = []
    wb = openpyxl.load_workbook(xlsx_path)
    profile = profiles.get_profile(profile_id or "rf-end")
    expected_cols = profile.headers
    is_rf_end = profile.id == "rf-end"

    # estrutura: Resumo primeiro, Leyenda último (só quando o perfil injeta essas abas)
    names = wb.sheetnames
    if profile.build_resumo and names and names[0] != "1. Resumo Executivo":
        warns.append(f"Primeira aba não é '1. Resumo Executivo' (é '{names[0]}').")
    if profile.build_leyenda and names and names[-1] != "Leyenda de Gaps":
        warns.append(f"Última aba não é 'Leyenda de Gaps' (é '{names[-1]}').")

    # colunas do perfil presentes; (rf-end) 8 antes das 5
    analyzable = []
    if extract_json:
        ex = json.loads(Path(extract_json).read_text(encoding="utf-8"))
        analyzable = [s for s in ex["sheets"] if s["analyzable"]]
    for s in analyzable:
        ws = wb[s["name"]] if s["name"] in wb.sheetnames else None
        if not ws:
            continue
        hr = s["header_row"]
        hdrs = [str(ws.cell(hr, c).value or "") for c in range(1, ws.max_column + 1)]
        for col in expected_cols:
            if col not in hdrs:
                errors.append(f"{s['name']}: coluna do perfil ausente no header: '{col}'.")
        if is_rf_end:
            try:
                i8 = hdrs.index(model.VALIDATION_COLUMNS[-1])
                i5 = hdrs.index(model.CONSULT_COLUMNS[0])
                if i5 < i8:
                    errors.append(f"{s['name']}: colunas consultivas antes das de validação.")
            except ValueError:
                pass

    # erros de fórmula + termos proibidos em todo o workbook
    formula_hits = 0
    forbidden_hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    v = cell.value
                    for fe in FORMULA_ERRORS:
                        if fe in v:
                            formula_hits += 1
                    low = v.lower()
                    for bad in FORBIDDEN_ANY:
                        if bad in low:
                            forbidden_hits.append(f"{ws.title}!{cell.coordinate}:'{bad.strip()}'")
    if formula_hits:
        errors.append(f"{formula_hits} célula(s) com erro de fórmula (#REF!/#NAME? etc.).")
    if forbidden_hits:
        errors.append(f"Termos proibidos no arquivo: {forbidden_hits[:8]}")
    return errors, warns


def run(analysis_json: str | None, xlsx: str | None, extract_json: str | None) -> int:
    all_err: list[str] = []
    all_warn: list[str] = []
    profile_id = None
    if analysis_json:
        data = json.loads(Path(analysis_json).read_text(encoding="utf-8"))
        profile_id = data.get("profile")
        e, w = validate_analysis(data)
        all_err += e
        all_warn += w
    if xlsx:
        e, w = validate_xlsx(xlsx, extract_json, profile_id)
        all_err += e
        all_warn += w

    print("=" * 60)
    if all_warn:
        print(f"WARNINGS ({len(all_warn)}):")
        for w in all_warn[:40]:
            print("  ⚠ ", w)
    if all_err:
        print(f"\nERROS ({len(all_err)}) — NÃO ENTREGAR:")
        for e in all_err[:60]:
            print("  ✗ ", e)
        print("\nGATE: REPROVADO")
        return 1
    print("GATE: APROVADO (sem erros bloqueantes)")
    return 0


def main() -> None:
    p = argparse.ArgumentParser(description="Valida a análise/entregável (estágio 4 do motor).")
    p.add_argument("analysis", nargs="?", default=None, help="analysis.json preenchido")
    p.add_argument("--xlsx", default=None, help="também valida o .xlsx de saída")
    p.add_argument("--extract", default=None, help="extract.json (para checar estrutura das colunas)")
    args = p.parse_args()
    if not args.analysis and not args.xlsx:
        p.error("informe analysis.json e/ou --xlsx")
    raise SystemExit(run(args.analysis, args.xlsx, args.extract))


if __name__ == "__main__":
    main()
