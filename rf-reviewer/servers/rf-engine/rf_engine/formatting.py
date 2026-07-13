"""
rf_engine.formatting — helpers determinísticos de estilo de planilha.

apply.py NÃO copia célula a célula para um workbook novo: abre o arquivo original
diretamente (openpyxl.load_workbook) e só ANEXA as colunas novas nos mesmos objetos
de planilha, salvando num caminho versionado. Como as células originais nunca são
recriadas, estilos/merges/hyperlinks/validações do cliente ficam intocados por
construção — não há necessidade de uma rotina de cópia separada.
"""
from __future__ import annotations

import math
import re

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import model


# ── helpers de estilo ────────────────────────────────────────────────────────
def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(bold: bool = False, color: str = "000000", size: int = 9, name: str = "Calibri") -> Font:
    return Font(bold=bold, color=color, size=size, name=name)


def align(h: str = "left", v: str = "center", wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def border_thin(hex_color: str = "BFBFBF") -> Border:
    s = Side(style="thin", color=hex_color)
    return Border(left=s, right=s, top=s, bottom=s)


# ── altura dinâmica de linha (rf-end-format-standard §4) ─────────────────────
def dynamic_row_height(values_widths: list[tuple[str, int]], minimum: int = 80) -> float:
    """Calcula altura da linha a partir do texto/largura das nossas colunas."""
    max_lines = 1
    for value, width in values_widths:
        text = "" if value is None else str(value)
        w = max(1, width)
        disp = sum(math.ceil(max(1, len(seg)) / w) for seg in text.split("\n"))
        max_lines = max(max_lines, disp)
    return max(minimum, max_lines * 14 + 10)


# ── scrub de termos proibidos (nunca vazar IA/ferramentas no entregável) ─────
def scrub_text(text: str) -> tuple[str, list[str]]:
    """Remove termos proibidos. Retorna (texto_limpo, termos_encontrados)."""
    if not text:
        return text, []
    found: list[str] = []
    out = text
    for term in model.FORBIDDEN_TERMS:
        pattern = re.compile(re.escape(term), re.IGNORECASE)
        if pattern.search(out):
            found.append(term)
            out = pattern.sub("", out)
    if found:
        out = re.sub(r"[ ]{2,}", " ", out).strip()
    return out, found
