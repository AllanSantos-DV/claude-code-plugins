"""
rf_engine.mcp_server — servidor MCP (stdio, JSON-RPC 2.0) do motor de revisão de RF.

Expõe o pipeline como TOOLS para o CrowdCode (claude-code-boss) plugar. O CrowdCode
cuida de sessão/controle; este MCP é AUTOCONTIDO: traz a disciplina do fluxo
(setup-antes → análise → injeção-depois → gate) sem depender do comportamento do
agente principal. Um agente-guia (AGENT.md) e as instruções (INSTRUCOES.md) acompanham.

Transporte: stdio, mensagens JSON-RPC 2.0 delimitadas por newline (padrão MCP stdio).
Sem dependências além de openpyxl (a mesma do motor). NADA vai pro stdout exceto o
protocolo — logs/avisos vão pro stderr.

Config no Claude Code (.mcp.json):
  { "mcpServers": { "rf-engine": {
      "command": "python", "args": ["-m", "rf_engine.mcp_server"],
      "cwd": "C:/caminho/para/rf-engine" } } }
"""
from __future__ import annotations

import contextlib
import io
import json
import sys
import traceback
from pathlib import Path

from . import __version__

PROTOCOL_VERSION = "2024-11-05"
SERVER_NAME = "rf-engine"


# ─────────────────────────────────────────────────────────────────────────────
# Implementação das tools (chamam os estágios do motor; capturam erros/stdout)
# ─────────────────────────────────────────────────────────────────────────────
def _guard(fn, **kw):
    """Executa um estágio capturando SystemExit e qualquer stdout acidental."""
    buf = io.StringIO()
    try:
        with contextlib.redirect_stdout(buf):
            return fn(**kw), None
    except SystemExit as exc:
        return None, str(exc.code if exc.code is not None else "erro")
    except Exception as exc:  # noqa: BLE001
        return None, f"{type(exc).__name__}: {exc}"


def tool_perfis_listar(_: dict) -> dict:
    from . import profiles
    return {"perfis": profiles.list_profiles()}


def tool_perfil_definir(args: dict) -> dict:
    from . import profiles
    spec = args.get("spec") or args
    p = profiles.define_profile(spec)
    return {"ok": True, "perfil": p.id, "colunas": p.headers}


def tool_prep(args: dict) -> dict:
    from . import extract as m_extract, scaffold as m_scaffold
    xlsx = args["xlsx"]
    out_dir = args.get("out_dir") or "_work/caso"
    perfil = args.get("perfil")
    ex, err = _guard(m_extract.run, xlsx_path=xlsx, out_dir=out_dir)
    if err:
        return {"ok": False, "stage": "extract", "error": err}
    analysis = str(Path(out_dir) / "analysis.json")
    sc, err = _guard(m_scaffold.run, extract_path=str(Path(out_dir) / "extract.json"),
                     out_path=analysis, profile_id=perfil)
    if err:
        return {"ok": False, "stage": "scaffold", "error": err}
    return {"ok": True, "out_dir": out_dir,
            "extract_json": str(Path(out_dir) / "extract.json"),
            "analysis_json": analysis,
            "perfil": sc.get("profile"), "perfil_colunas": sc.get("profile_columns"),
            "analyzable_sheets": ex["analyzable_sheets"], "counts": sc["counts"],
            "schema_keys": sc["schema_keys"],
            "next": "AGENTE preenche as schema_keys + 'gaps' em analysis_json, depois chame rf_apply."}


def tool_brain_buscar(args: dict) -> dict:
    from . import brain_client as m_brain
    c = m_brain.BrainClient(args["url"], args.get("project", "la-positiva"))
    try:
        c.connect()
        hits = c.search(args["query"], top_k=int(args.get("top_k", 5)))
    finally:
        c.close()
    return {"ok": True, "hits": hits}


def tool_brain_enriquecer(args: dict) -> dict:
    from . import brain_client as m_brain
    res, err = _guard(m_brain.enrich, analysis_path=args["analysis_json"], url=args["url"],
                      project=args.get("project", "la-positiva"), top_k=int(args.get("top_k", 5)),
                      limit=args.get("limit"))
    if err:
        return {"ok": False, "error": err}
    return {"ok": True, **res}


def tool_apply(args: dict) -> dict:
    from . import apply as m_apply
    res, err = _guard(m_apply.run, extract_json=args["extract_json"], analysis_json=args["analysis_json"],
                      out_dir=args.get("out_dir"), profile_id=args.get("perfil"))
    if err:
        return {"ok": False, "error": err}
    return {"ok": True, **res}


def tool_validar(args: dict) -> dict:
    from . import validate as m_validate
    analysis = args.get("analysis_json")
    xlsx = args.get("xlsx")
    extract = args.get("extract_json")
    errors: list[str] = []
    warns: list[str] = []
    if analysis:
        data = json.loads(Path(analysis).read_text(encoding="utf-8"))
        e, w = m_validate.validate_analysis(data)
        errors += e
        warns += w
        pid = data.get("profile")
    else:
        pid = None
    if xlsx:
        e, w = m_validate.validate_xlsx(xlsx, extract, pid)
        errors += e
        warns += w
    return {"ok": True, "aprovado": len(errors) == 0, "errors": errors[:60], "warnings": warns[:40]}


def tool_verificar_preservacao(args: dict) -> dict:
    """Prova mecânica: base × saída, célula a célula (write-back não-destrutivo)."""
    import openpyxl
    import warnings as _w
    _w.filterwarnings("ignore")
    base = openpyxl.load_workbook(args["base_xlsx"], data_only=False)
    out = openpyxl.load_workbook(args["out_xlsx"], data_only=False)

    def norm(v):
        t = getattr(v, "text", None)
        return str(t) if t is not None else ("" if v is None else str(v))

    added = [s for s in out.sheetnames if s not in base.sheetnames]
    removed = [s for s in base.sheetnames if s not in out.sheetnames]
    total = mismatch = 0
    examples = []
    for name in base.sheetnames:
        if name not in out.sheetnames:
            continue
        wsB, wsO = base[name], out[name]
        for r in range(1, (wsB.max_row or 0) + 1):
            for c in range(1, (wsB.max_column or 0) + 1):
                total += 1
                if norm(wsB.cell(r, c).value) != norm(wsO.cell(r, c).value):
                    mismatch += 1
                    if len(examples) < 15:
                        examples.append(f"{name}!R{r}C{c}")
    return {"ok": mismatch == 0 and not removed, "celulas_originais": total,
            "divergencias": mismatch, "abas_novas": added, "abas_removidas": removed,
            "exemplos_divergencia": examples,
            "veredito": "MECANICO E NAO-DESTRUTIVO" if (mismatch == 0 and not removed) else "REVISAR"}


def tool_status(args: dict) -> dict:
    d = Path(args.get("out_dir") or "_work/caso")
    ex = d / "extract.json"
    an = d / "analysis.json"
    outs = list((d / "out").glob("*.xlsx")) if (d / "out").exists() else []
    filled = None
    perfil = None
    if an.exists():
        data = json.loads(an.read_text(encoding="utf-8"))
        perfil = data.get("profile")
        keys = data.get("schema_keys", [])
        active = [r for r in data.get("records", []) if not r.get("na")]
        done = sum(1 for r in active if all((r.get(k) or "").strip() for k in keys))
        filled = {"preenchidos": done, "total_ativos": len(active),
                  "completo": done == len(active) and len(active) > 0}
    stage = ("apply-feito" if outs else
             "analise-preenchida" if (filled and filled["completo"]) else
             "aguardando-analise" if an.exists() else
             "extraido" if ex.exists() else "vazio")
    return {"ok": True, "out_dir": str(d), "stage": stage, "perfil": perfil,
            "extract": ex.exists(), "analysis": an.exists(), "preenchimento": filled,
            "saidas": [str(p) for p in outs]}


TOOLS = {
    "rf_perfis_listar": (tool_perfis_listar, "Lista os perfis de coluna (o molde da saída). Sem argumentos.", {
        "type": "object", "properties": {}}),
    "rf_perfil_definir": (tool_perfil_definir, "Registra/atualiza um perfil custom de colunas (persiste).", {
        "type": "object", "properties": {
            "spec": {"type": "object", "description": "id, name, lang, columns[{header,key,width,header_fill,enum,colored}], build_resumo, build_leyenda"}},
        "required": ["spec"]}),
    "rf_prep": (tool_prep, "SETUP ANTES da análise: extrai a planilha e monta o esqueleto (extract+scaffold). Escolhe o perfil de colunas.", {
        "type": "object", "properties": {
            "xlsx": {"type": "string", "description": "caminho do entregável .xlsx do cliente"},
            "out_dir": {"type": "string", "description": "pasta de trabalho (default _work/caso)"},
            "perfil": {"type": "string", "description": "id do perfil (default rf-end; ex.: fernando-siniestros)"}},
        "required": ["xlsx"]}),
    "rf_brain_buscar": (tool_brain_buscar, "Busca evidências no cérebro (servidor de memória) escopado por project_id. Retrieval mecânico.", {
        "type": "object", "properties": {
            "url": {"type": "string", "description": "serverUrl (ex.: http://192.168.18.13:38080)"},
            "project": {"type": "string", "description": "project_id (default la-positiva)"},
            "query": {"type": "string"}, "top_k": {"type": "integer"}},
        "required": ["url", "query"]}),
    "rf_brain_enriquecer": (tool_brain_enriquecer, "Enriquece o analysis.json com candidate_refs do cérebro (1 busca por RF).", {
        "type": "object", "properties": {
            "analysis_json": {"type": "string"}, "url": {"type": "string"},
            "project": {"type": "string"}, "top_k": {"type": "integer"}, "limit": {"type": "integer"}},
        "required": ["analysis_json", "url"]}),
    "rf_apply": (tool_apply, "INJEÇÃO DEPOIS da análise: devolve a análise aprovada pra dentro da planilha (não-destrutivo, versiona). Usa o perfil.", {
        "type": "object", "properties": {
            "extract_json": {"type": "string"}, "analysis_json": {"type": "string"},
            "out_dir": {"type": "string"}, "perfil": {"type": "string"}},
        "required": ["extract_json", "analysis_json"]}),
    "rf_validar": (tool_validar, "GATE de qualidade: valida a análise e/ou o .xlsx (anti-padrões, enums, colunas). aprovado=true libera entrega.", {
        "type": "object", "properties": {
            "analysis_json": {"type": "string"}, "xlsx": {"type": "string"}, "extract_json": {"type": "string"}}}),
    "rf_verificar_preservacao": (tool_verificar_preservacao, "Prova que a injeção não alterou o original: compara base × saída célula a célula.", {
        "type": "object", "properties": {
            "base_xlsx": {"type": "string"}, "out_xlsx": {"type": "string"}},
        "required": ["base_xlsx", "out_xlsx"]}),
    "rf_status": (tool_status, "Estado do pipeline numa pasta de trabalho: extraído / aguardando análise / preenchido / apply-feito.", {
        "type": "object", "properties": {"out_dir": {"type": "string"}}}),
}


# ─────────────────────────────────────────────────────────────────────────────
# Loop JSON-RPC 2.0 sobre stdio
# ─────────────────────────────────────────────────────────────────────────────
def _send(msg: dict) -> None:
    sys.stdout.write(json.dumps(msg, ensure_ascii=False) + "\n")
    sys.stdout.flush()


def _result(rid, result) -> None:
    _send({"jsonrpc": "2.0", "id": rid, "result": result})


def _error(rid, code: int, message: str) -> None:
    _send({"jsonrpc": "2.0", "id": rid, "error": {"code": code, "message": message}})


def _handle(msg: dict) -> None:
    method = msg.get("method")
    rid = msg.get("id")
    if method == "initialize":
        _result(rid, {
            "protocolVersion": PROTOCOL_VERSION,
            "capabilities": {"tools": {}},
            "serverInfo": {"name": SERVER_NAME, "version": __version__},
        })
    elif method == "notifications/initialized":
        return  # notificação, sem resposta
    elif method == "ping":
        _result(rid, {})
    elif method == "tools/list":
        tools = [{"name": n, "description": d, "inputSchema": s} for n, (_, d, s) in TOOLS.items()]
        _result(rid, {"tools": tools})
    elif method == "tools/call":
        params = msg.get("params") or {}
        name = params.get("name")
        args = params.get("arguments") or {}
        entry = TOOLS.get(name)
        if not entry:
            _error(rid, -32601, f"tool desconhecida: {name}")
            return
        fn = entry[0]
        try:
            out = fn(args)
            payload = json.dumps(out, ensure_ascii=False, indent=2)
            is_err = isinstance(out, dict) and out.get("ok") is False
            _result(rid, {"content": [{"type": "text", "text": payload}], "isError": is_err})
        except Exception as exc:  # noqa: BLE001
            tb = traceback.format_exc()
            sys.stderr.write(tb + "\n")
            _result(rid, {"content": [{"type": "text", "text": f"ERRO: {type(exc).__name__}: {exc}"}],
                          "isError": True})
    elif rid is not None:
        _error(rid, -32601, f"método não suportado: {method}")


def main() -> None:
    sys.stderr.write(f"[rf-engine mcp] iniciado v{__version__} ({len(TOOLS)} tools)\n")
    sys.stderr.flush()
    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue
        try:
            msg = json.loads(line)
        except json.JSONDecodeError:
            continue
        try:
            _handle(msg)
        except Exception:  # noqa: BLE001
            sys.stderr.write(traceback.format_exc() + "\n")
            sys.stderr.flush()


if __name__ == "__main__":
    main()
